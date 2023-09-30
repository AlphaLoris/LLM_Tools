import os
import tkinter as tk
from tkinter import filedialog
from tkinter import ttk
from tkinter.scrolledtext import ScrolledText
from tkinter import messagebox
from tkinter import simpledialog
import tiktoken
import openai
import json
import textwrap
import keyring.backends.macOS
import keyring.backends.Windows
from jsonschema import validate, ValidationError
import datetime
import pandas as pd
from openpyxl import load_workbook
import chardet

"""
To build this script into an executable, run the following command from the root directory of the project:
pyinstaller --hidden-import=tiktoken_ext.openai_public --hidden-import=tiktoken_ext --onefile openai_chat_gui.py
"""


def check_api_key(root, keychain_path=None):
    # Attempt to retrieve the API key from the keyring
    print("Attempting to retrieve API Key from keyring")
    api_key = keyring.get_password("openai", "api_key")
    if not api_key:
        # Open a Tkinter dialogue to prompt the user for their OpenAI API Key
        print("API Key not retrieved from keyring. Prompting user for API Key")
        api_key = get_api_key(root, keychain_path)
    return api_key


def get_api_key(root, keychain_path=None):
    api_key = prompt_for_api_key(root)

    # Validate the API key
    print("User entered API Key. Validating...")
    test_model = "gpt-3.5-turbo"
    errors = is_valid_api_key_model(api_key, test_model)

    if errors:
        # Notify the user in the Tkinter dialogue and allow them to correct the API Key
        print("API Key is invalid. Prompting user to correct API Key")
        notify_invalid_key(errors, api_key)
        return check_api_key(root, keychain_path)

    # Write the API key to the keyring
    print("API Key is valid. Writing to keyring and returning it.", api_key)
    keyring.set_password("openai", "api_key", api_key)

    return api_key


class CustomDialog(simpledialog.Dialog):
    def __init__(self, parent, title=None):
        self.entry = None
        self.result = None
        super().__init__(parent, title=title)

    def body(self, parent):
        tk.Label(parent, text="Enter your OpenAI API Key:").grid(row=0)
        self.entry = tk.Entry(parent, width=75)
        self.entry.grid(row=1, padx=10, pady=10)
        return self.entry

    def apply(self):
        self.result = self.entry.get()


def prompt_for_api_key(root):
    print("Opening dialog to prompt user for API Key")
    dialog = CustomDialog(root, "OpenAI API Key")
    api_key = dialog.result.strip()
    return api_key


def is_valid_api_key_model(api_key, test_model):
    openai.api_key = api_key
    error_messages = []
    print("Validating API key by calling OpenAI API")
    try:
        response = openai.ChatCompletion.create(
            model=test_model,
            messages=[{"role": "user", "content": "Hello, world!"}],
            temperature=0.9, top_p=1, n=1, stream=False, max_tokens=5,
            presence_penalty=0, frequency_penalty=0, logit_bias={}, user=""
        )
        print("This API key/model combination is valid. Model Response:")
        print(response['choices'][0]['message']['content'])
    except openai.OpenAIError as e:
        print(f"Error: {e}")
        error_messages.append(str(e))
    return error_messages


# TODO: This function displays an error message that seems to be irrelevant to this application. Need to fix it.
def notify_invalid_key(errors, api_key):
    print("OpenAI API call failed. Notifying user that API Key is invalid")
    error_message = "\n".join(errors)
    message = f"Invalid API Key: {api_key}\n{error_message}"
    messagebox.showerror("Invalid API Key", message)


class CustomMessageBox(simpledialog.Dialog):
    def __init__(self, parent, title, message):
        self.result = None
        self.message = message
        super().__init__(parent, title=title)

    def body(self, master):
        tk.Label(master, text=self.message).pack()
        return None

    def buttonbox(self):
        box = tk.Frame(self)

        tk.Button(
            box, text="Continue", width=10, command=self.continue_action
        ).pack(side=tk.LEFT, padx=5, pady=5)
        tk.Button(
            box, text="Return", width=10, command=self.go_back_action
        ).pack(side=tk.LEFT, padx=5, pady=5)

        self.bind("<Return>", self.continue_action)
        self.bind("<Escape>", self.go_back_action)

        box.pack()

    def continue_action(self, event=None):
        self.result = True
        self.destroy()

    def go_back_action(self, event=None):
        self.result = False
        self.destroy()


# Function to determine the number of tokens in a message
def num_tokens_from_messages(messages, model):
    """Return the number of tokens used by a list of messages."""
    try:
        encoding = tiktoken.encoding_for_model(model)
    except KeyError:
        print("Warning: model not found. Using cl100k_base encoding.")
        encoding = tiktoken.get_encoding("cl100k_base")
    if model in {
        "gpt-3.5-turbo-0613",
        "gpt-3.5-turbo-16k-0613",
        "gpt-4-0314",
        "gpt-4-32k-0314",
        "gpt-4-0613",
        "gpt-4-32k-0613",
    }:
        tokens_per_message = 3
        tokens_per_name = 1
    elif model == "gpt-3.5-turbo-0301":
        tokens_per_message = 4  # every message follows <|start|>{role/name}\n{content}<|end|>\n
        tokens_per_name = -1  # if there's a name, the role is omitted
    elif "gpt-3.5-turbo" in model:
        print("Warning: gpt-3.5-turbo may update over time. Returning num tokens assuming gpt-3.5-turbo-0613.")
        return num_tokens_from_messages(messages, model="gpt-3.5-turbo-0613")
    elif "gpt-4" in model:
        print("Warning: gpt-4 may update over time. Returning num tokens assuming gpt-4-0613.")
        return num_tokens_from_messages(messages, model="gpt-4-0613")
    else:
        raise NotImplementedError(
            f"""num_tokens_from_messages() is not implemented for model {model}
            . See https://github.com/openai/openai-python/blob/main/chatml.md for information on how messages are
             converted to tokens."""
        )

    # num_tokens += len(encoding.encode(message, disallowed_special=()))
    num_tokens = 0
    for message in messages:
        num_tokens += tokens_per_message
        for key, value in message.items():
            if isinstance(value, str):
                # Encode strings
                num_tokens += len(encoding.encode(value, disallowed_special=()))
            elif isinstance(value, (int, float)):
                # Convert numbers to strings
                num_tokens += len(encoding.encode(str(value)))
            elif isinstance(value, (list, dict)):
                # Handle iterables separately
                num_tokens += len(value)
            else:
                # For other types, try converting to string
                try:
                    str_value = str(value)
                    num_tokens += len(encoding.encode(str_value, disallowed_special=()))
                except:
                    print(f"Could not handle value: {value}")
    num_tokens += 3  # every reply is primed with <|start|>assistant<|message|>
    return num_tokens


def send_request(model, prompt, temperature, top_p, n, stream, stop, max_tokens, presence_penalty, frequency_penalty, logit_bias, user):

    # Ensure OpenAI library is imported
    import openai

    # Helper function to extract and print StringVar parameters
    def extract_string_var(value, name):
        if isinstance(value, tk.StringVar):
            print(f"'{name}' was a StringVar. Extracting its value.")
            return value.get()
        return value

    model = extract_string_var(model, "model")
    temperature = extract_string_var(temperature, "temperature")
    top_p = extract_string_var(top_p, "top_p")
    n = extract_string_var(n, "n")
    max_tokens = extract_string_var(max_tokens, "max_tokens")
    presence_penalty = extract_string_var(presence_penalty, "presence_penalty")
    frequency_penalty = extract_string_var(frequency_penalty, "frequency_penalty")
    user = extract_string_var(user, "user")

    # For lists like 'stop' and 'logit_bias', ensure they are extracted properly if they are in StringVar
    stop = [extract_string_var(s, "stop item") for s in stop]
    logit_bias = {k: extract_string_var(v, f"logit_bias for token {k}") for k, v in logit_bias.items()}

    # Prepare the payload
    payload = {
        "model": model,
        "messages": prompt,
        "temperature": temperature,
        "top_p": top_p,
        "n": n,
        "stream": stream,
        "stop": stop,
        "max_tokens": max_tokens,
        "presence_penalty": presence_penalty,
        "frequency_penalty": frequency_penalty,
        "logit_bias": logit_bias,
        "user": user
    }

    print("Sending request to OpenAI in send_request function...")

    # Use the OpenAI library to send the request
    try:
        response = openai.ChatCompletion.create(**payload)
        return response
    except openai.OpenAIError as e:
        print(f"Error while sending request to OpenAI: {e}")
        return None


# Main UI
class ResultsTab(tk.Frame):
    def __init__(self, parent):
        super().__init__(parent)

        label = tk.Label(self, text="Response")
        label.grid(row=0, column=0, pady=10)

        prompts_listbox = tk.Listbox(self)
        prompts_listbox.grid(row=1, column=0, padx=10, pady=10, sticky="nsew")

        self.columnconfigure(0, weight=1)
        self.rowconfigure(1, weight=1)


class ToolTip:
    def __init__(self, widget, text):
        self.widget = widget
        self.text = text
        self.tooltip_window = None
        self.widget.bind("<Enter>", self.show_tooltip)
        self.widget.bind("<Leave>", self.hide_tooltip)

    def show_tooltip(self, event=None):
        x, y, _, _ = self.widget.bbox("insert")
        x += self.widget.winfo_rootx() + 25
        y += self.widget.winfo_rooty() + 20
        self.tooltip_window = tk.Toplevel(self.widget)
        self.tooltip_window.wm_overrideredirect(True)
        self.tooltip_window.wm_geometry(f"+{x}+{y}")
        label = tk.Label(self.tooltip_window, text=self.text, background="light grey", relief="solid", borderwidth=1)
        label.pack()

    def hide_tooltip(self, event=None):
        if self.tooltip_window:
            self.tooltip_window.destroy()
            self.tooltip_window = None


class PromptUI:
    def __init__(self, root, model_list, context_windows, api_key):
        self.root = root
        self.model_list = model_list
        self.api_key = api_key
        self.context_windows = context_windows
        self.context_length = tk.IntVar()
        self.root.title("Chat Completion API UI")
        self.root.geometry("700x800")

        self.prompt_notebook = ttk.Notebook(self.root)
        self.prompt_notebook.grid(row=0, column=0, sticky="nsew")

        self.prompt_tab = tk.Frame(self.prompt_notebook)
        self.prompt_notebook.add(self.prompt_tab, text=" Prompt ")

        self.results_tab = ResultsTab(self.prompt_notebook)
        self.prompt_notebook.add(self.results_tab, text=" Response ")

        self.prompt_test_tab = tk.Frame(self.prompt_notebook)
        self.prompt_notebook.add(self.prompt_test_tab, text=" Prompt Test Resources ")

        self.prompt_structure_tab = tk.Frame(self.prompt_notebook)
        self.prompt_notebook.add(self.prompt_structure_tab, text=" Guide to Prompt Structure ")

        prompt_structure_text = textwrap.dedent("""Context Window Size (in tokens):

gpt-4: 8192
gpt-4-0613: 8192
gpt-4-0314: 8192
gpt-4-32k: 32768
gpt-4-32k-0613: 32768
gpt-4-32k-0314: 32768
gpt-3.5-turbo: 4096
gpt-3.5-turbo-16k: 16384
gpt-3.5-turbo-0613: 4096
gpt-3.5-turbo-0301: 4096

Guide to Structuring a Chat Completion prompt for OpenAI LLMs

Note: the maximum length of a prompt is limited by the context window size. This is the sum of all of the tokens in the message components, plus the tokens that OpenAI adds to the prompt, and the tokens in the response.

Chat completion prompts are composed of as many as three types of messages: "System" messages, "Assistant" messages, and "User" messages. These elements serve the following purposes:

- The "System" message provides the LLM guidance on the expectations for its behavior in responding to prompts.  A typical system message might be "You are an extremely helpful AI Python coding assistant with deep experience building complex Tkinter UIs"

- The "Assistant" messages are used to model the type of output the prompter expects from the model. This is primary for use in few-shot prompting.

- The "User" message is the primary component, and is frequently used by itself or in combination with a "System" message. This message conveys the intent of the prompter to the model, providing context, a task, and instructions as to how to complete the task or constraints to shape the response.

A typical prompt might be structured as follows:

"System: You are an AI assistant with access to all of human wisdom who is eager to share your knowledge with humans seeking that wisdom.

User: Your task is to answer my questions in a consistent style that follows the example below.

User: Teach me about patience.  

Assistant: The river that carves the deepest valley flows from a modest spring; the grandest symphony originates from a single note; the most intricate tapestry begins with a solitary thread.

User: Teach me about resilience."

This prompt would be built in this GUI by creating message components for each of the messages above, and then assigning a message type/role to each of the message components.  The model might respond with the following response:

Resilience is like a tree that bends with the wind but never breaks. It is the ability to bounce back from adversity and keep moving forward, even when things get tough. Just like a tree that grows stronger with each storm it weathers, resilience is a quality that can be developed and strengthened over time.

This example is paraphrased from the OpenAI ChatGPT Prompt Engineering for Developers course on deeplearning.ai:

https://www.deeplearning.ai/short-courses/chatgpt-prompt-engineering-for-developers/

Further guidance on prompting can be found here: 

https://www.promptingguide.ai/
        """)

        scrollbar = tk.Scrollbar(self.prompt_structure_tab)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        prompt_structure_textbox = tk.Text(self.prompt_structure_tab, wrap=tk.WORD, height=10, width=50)
        prompt_structure_textbox.insert(tk.END, prompt_structure_text)
        prompt_structure_textbox.configure(state='disabled')
        scrollbar.config(command=prompt_structure_textbox.yview)
        prompt_structure_textbox.pack(padx=10, pady=10, fill=tk.BOTH, expand=True)

        self.root.columnconfigure(0, weight=1)
        self.root.rowconfigure(0, weight=1)

        self.content_frame = tk.Frame(self.prompt_tab)
        self.content_frame.grid(row=0, column=0, sticky="nsew", rowspan=2)

        # Parameters frame
        self.parameters_frame = tk.Frame(self.content_frame)
        self.parameters_frame.grid(row=0, column=0, rowspan=3, columnspan=3, sticky="w")
        self.parameters_frame.columnconfigure(0, weight=0, minsize=5)
        self.parameters_frame.columnconfigure(1, weight=0, minsize=5)
        self.parameters_frame.columnconfigure(2, weight=0, minsize=5)
        self.parameters_frame.columnconfigure(3, weight=0, minsize=5)
        self.parameters_frame.columnconfigure(4, weight=0, minsize=5)
        self.parameters_frame.columnconfigure(5, weight=1)

        # Model
        self.model_var = tk.StringVar(self.parameters_frame)
        self.model_var.trace_add('write', self.update_context_length)
        self.refresh_button = tk.Button(self.parameters_frame, text="Refresh Models",
                                        command=self.refresh_model_list)
        print("Model_list: ", model_list)
        self.refresh_button.grid(row=0, column=0, padx=10, pady=5, sticky="w")
        self.model_var.set("Select Model")
        self.model_menu = ttk.Combobox(self.parameters_frame, textvariable=self.model_var)
        self.model_menu['values'] = ["Select Model"] + model_list
        self.model_menu.set("Select Model")  # default value
        self.model_menu.grid(row=0, column=1, padx=10, pady=5, sticky="w", columnspan=1)

        # API Key
        self.api_key_entry = tk.Entry(self.parameters_frame, width=85)
        self.api_key_entry.grid(row=1, column=1, padx=5, columnspan=4)
        self.api_key_entry.insert(0, openai.api_key)
        self.edit_button = tk.Button(self.parameters_frame, text="Edit API Key", command=self.edit_api_key)
        self.edit_button.grid(row=1, column=0, padx=10)
        self.label_api_key = tk.Label(self.parameters_frame, text="API Key", width=13, anchor="w")
        self.label_api_key.grid(row=1, column=5, padx=3)
        api_key_tooltip_text = "The API Key is the code necessary to access the API of the LLM.\n " \
                               "The API Key For access to the OpenAI Chat Completion API is available on\n " \
                               "your account page on the OpenAI website at this URL:\n\n" \
                               "https://platform.openai.com/account/api-keys\n\n" \
                               "(You may have to associate a payment method with your account to generate an API Key.)"
        ToolTip(self.label_api_key, api_key_tooltip_text)

        # Temperature
        self.temperature_entry = tk.Entry(self.parameters_frame, width=5)
        self.temperature_entry.grid(row=2, column=0, padx=10, pady=5, sticky="nswe")
        self.temperature_entry.insert(0, "0.7")
        self.temperature_label = tk.Label(self.parameters_frame, text="temperature (0-2)", width=13, anchor="w")
        self.temperature_label.grid(row=2, column=1, padx=5, pady=5, sticky="w")
        temperature_tooltip_text = "The temperature parameter controls the randomness of the model's response\n" \
                                   "by controlling the likelihood that the model will return the most probable \n" \
                                   "next token (word) versus a less probable next token. The higher the" \
                                   " temperature, \nthe more random the choice of next token."
        ToolTip(self.temperature_label, temperature_tooltip_text)

        # Top_p
        self.top_p_entry = tk.Entry(self.parameters_frame, width=5)
        self.top_p_entry.grid(row=3, column=0, padx=10, pady=5, sticky="nswe")
        self.top_p_entry.insert(0, "1.0")
        self.top_p_label = tk.Label(self.parameters_frame, text="top_p (0-1)", width=9, anchor="w")
        self.top_p_label.grid(row=3, column=1, padx=5, pady=5, sticky="w")
        top_p_tooltip_text = "The top_p parameter also controls the randomness of the model's response \n" \
                             "by limiting the set of tokens that the model is allowed to select from when\n" \
                             "choosing the next token.  If top_p is set to 0.1, the model must choose the\n" \
                             "next token (word) from the top 10% most probable tokens. The higher the\n" \
                             "top_p value,the larger the set of tokens the model has to choose from."
        ToolTip(self.top_p_label, top_p_tooltip_text)

        # n
        self.n_entry = tk.Entry(self.parameters_frame, width=10)
        self.n_entry.grid(row=4, column=0, padx=10, pady=5, sticky="nswe")
        self.n_entry.insert(0, "1")
        self.n_label = tk.Label(self.parameters_frame, text="n (1-4)", width=9, anchor="w")
        self.n_label.grid(row=4, column=1, padx=5, pady=5, sticky="w")
        n_label_tooltip_text = "The n parameter determines the number of different completions the model will \n" \
                               "return.  If n is set to 3, the model will return three completions."
        ToolTip(self.n_label, n_label_tooltip_text)

        # stream - not implemented

        # stop
        self.stop_entry = tk.Entry(self.parameters_frame, width=5)
        self.stop_entry.grid(row=5, column=0, padx=10, pady=5, sticky="nswe")
        self.stop_entry.insert(0, "")
        self.stop_label = tk.Label(self.parameters_frame, text="stop (text string(s))", width=14, anchor="w")
        self.stop_label.grid(row=5, column=1, padx=5, pady=5, sticky="w")
        stop_label_tooltip_text = 'The stop parameter determines when the model will stop generating text. If the \n' \
                                  'model generates one of the strings (you can provide up to four strings) provided, it \n' \
                                  'will stop generating text and return the result, which will not include the stop \n' \
                                  'string. The strings should be enclosed in quotation marks and separated by a comma\n' \
                                  'and a space. For example, "11.", "Hello Michael", "Go home!"'
        ToolTip(self.stop_label, stop_label_tooltip_text)

        # max_tokens
        self.max_tokens_entry = tk.Entry(self.parameters_frame, width=5)
        self.max_tokens_entry.grid(row=6, column=0, padx=10, pady=5, sticky="nswe")
        self.max_tokens_entry.insert(0, "")
        self.max_tokens_label = tk.Label(self.parameters_frame, text="max_tokens (1-?)", width=16, anchor="w")
        self.max_tokens_label.grid(row=6, column=1, padx=5, pady=5, sticky="w")
        max_tokens_label_tooltip_text = "The max_tokens parameter specifies the maximum number of tokens the model is\n" \
                                        "to generate in the chat completion. The total length of input tokens and\n" \
                                        "generated tokens is limited by the model's context window. For GPT-4,\n" \
                                        "the context length is 8192 tokens, So the max_tokens value is limited to\n" \
                                        "the context length - prompt message tokens, but the value you assign can\n" \
                                        "be less than that."
        ToolTip(self.max_tokens_label, max_tokens_label_tooltip_text)

        # presence_penalty
        self.presence_penalty_entry = tk.Entry(self.parameters_frame, width=5)
        self.presence_penalty_entry.grid(row=2, column=2, padx=5, pady=5, sticky="nswe")
        self.presence_penalty_entry.insert(0, "")
        self.presence_penalty_label = tk.Label(self.parameters_frame, text="presence_penalty (-2.0 to 2.0)", width=24,
                                               anchor="w")
        self.presence_penalty_label.grid(row=2, column=3, padx=5, pady=5, sticky="w")
        presence_penalty_label_tooltip_text = "Presence_penalty controls the model's likelihood of using words that\n" \
                                              "already occur in its output again. Positive values penalize new tokens \n" \
                                              "based on whether they appear in the text so far, increasing the model's\n" \
                                              "likelihood to talk about new topics."
        ToolTip(self.presence_penalty_label, presence_penalty_label_tooltip_text)

        # frequency_penalty
        self.frequency_penalty_entry = tk.Entry(self.parameters_frame, width=5)
        self.frequency_penalty_entry.grid(row=3, column=2, padx=5, pady=5, sticky="nswe")
        self.frequency_penalty_entry.insert(0, "")
        self.frequency_penalty_label = tk.Label(self.parameters_frame, text="frequency_penalty (-2.0 to 2.0)", width=24,
                                                anchor="w")
        self.frequency_penalty_label.grid(row=3, column=3, padx=5, pady=5, sticky="w")
        frequency_penalty_label_tooltip_text = "Frequency_penalty is a number between -2.0 and 2.0 that defaults to 0.\n " \
                                               "Positive values penalize new tokens based on their existing frequency in\n" \
                                               "the text so far, decreasing the model's likelihood to repeat the same \n" \
                                               "line verbatim."
        ToolTip(self.frequency_penalty_label, frequency_penalty_label_tooltip_text)

        # logit_bias
        self.logit_bias_entry = tk.Entry(self.parameters_frame, width=5)
        self.logit_bias_entry.grid(row=4, column=2, padx=5, pady=5, sticky="nswe")
        self.logit_bias_entry.insert(0, "")
        self.logit_bias_label = tk.Label(self.parameters_frame, text="logit_bias (JSON)", width=16,
                                         anchor="w")
        self.logit_bias_label.grid(row=4, column=3, padx=5, pady=5, sticky="w")
        logit_bias_label_tooltip_text = "The logit_bias parameter is a JSON object that allows you to adjust the\n" \
                                        "likelihood of the model generating certain words. The keys are tokens and\n" \
                                        "the values are bias values. Positive bias values increase the likelihood of\n" \
                                        "the model generating the associated word, while negative values decrease the\n" \
                                        "likelihood. The bias values are on a log scale, so a value of 1.0 increases\n" \
                                        "the likelihood by 10x, and a value of -1.0 decreases the likelihood by 10x."
        ToolTip(self.logit_bias_label, logit_bias_label_tooltip_text)

        # user
        self.user_entry = tk.Entry(self.parameters_frame, width=5)
        self.user_entry.grid(row=5, column=2, padx=5, pady=5, sticky="nswe")
        self.user_entry.insert(0, "")
        self.user_label = tk.Label(self.parameters_frame, text="user [user ID]", width=16,
                                   anchor="w")
        self.user_label.grid(row=5, column=3, padx=5, pady=5, sticky="w")
        user_label_tooltip_text = "The user parameter is a string that allows you to associate a prompt with a\n" \
                                  "specific user. This is intended to be used for tracking which user within an\n" \
                                  " organization submitted which prompts."
        ToolTip(self.user_label, user_label_tooltip_text)

        # Context length
        self.context_length = tk.IntVar()
        self.context_length.set(0)
        self.context_length_value = tk.Label(self.parameters_frame, textvariable=self.context_length, anchor="w")
        self.context_length_value.grid(row=11, column=0, padx=10, pady=5, sticky="w")
        self.context_length_label = tk.Label(self.parameters_frame, text="Context length",
                                             width=15, anchor="w")
        self.context_length_label.grid(row=11, column=1, columnspan=2, pady=5, sticky="w")
        context_length_label_tooltip_text = "This is the context length of the selected model in tokens."
        ToolTip(self.context_length_label, context_length_label_tooltip_text)

        # Token count
        self.token_count = tk.IntVar()
        self.token_count.set(0)
        self.token_count_value = tk.Label(self.parameters_frame, textvariable=self.token_count, anchor="w")
        self.token_count_value.grid(row=12, column=0, padx=10, pady=5, sticky="w")
        self.token_count_label = tk.Label(self.parameters_frame, text="Current token count", width=15, anchor="w")
        self.token_count_label.grid(row=12, column=1, columnspan=2, pady=5, sticky="w")
        token_count_label_tooltip_text = "This is the total token count for all of the message components that have\n" \
                                         "been created.\n"
        ToolTip(self.token_count_label, token_count_label_tooltip_text)

        # Message buttons frame
        self.message_buttons_frame = tk.Frame(self.content_frame)
        # Assuming content_frame is your parent frame
        self.content_frame.columnconfigure(0, weight=1)

        # Message buttons frame
        self.message_buttons_frame = tk.Frame(self.content_frame)
        self.message_buttons_frame.grid(row=4, column=0, rowspan=1, columnspan=4)

        self.message_buttons_frame.columnconfigure(0, weight=1)  # Empty column on the left
        self.message_buttons_frame.columnconfigure(1, weight=0)  # Button 1
        self.message_buttons_frame.columnconfigure(2, weight=0)  # Button 2
        self.message_buttons_frame.columnconfigure(3, weight=0)  # Button 3
        self.message_buttons_frame.columnconfigure(4, weight=0)  # Button 3
        self.message_buttons_frame.columnconfigure(5, weight=1)  # Empty column on the right

        # Buttons
        self.new_prompt_button = tk.Button(self.message_buttons_frame, text="Restore Defaults",
                                           command=self.restore_default_values)
        self.new_prompt_button.grid(row=3, column=1, padx=5, pady=5)

        self.add_message_component_button = tk.Button(self.message_buttons_frame, text="Add message component",
                                                      command=self.open_message_component_window)
        self.add_message_component_button.grid(row=3, column=2, padx=5, pady=5)

        self.submit_prompt_button = tk.Button(self.message_buttons_frame, text="Submit prompt",
                                              command=self.submit_prompt)
        self.submit_prompt_button.grid(row=3, column=3, padx=5, pady=5)

        self.submit_prompt_button = tk.Button(self.message_buttons_frame, text="Delete prompt",
                                              command=self.delete_prompt)
        self.submit_prompt_button.grid(row=3, column=4, padx=5, pady=5)

        # Message components frame
        self.message_components = []

        self.message_components_frame = tk.Frame(self.content_frame)
        self.message_components_frame.grid(row=5, column=0, columnspan=3, sticky="w")

        self.message_components_frame.columnconfigure(0, weight=1, minsize=5)
        self.message_components_frame.columnconfigure(1, weight=0, minsize=20)
        self.message_components_frame.columnconfigure(2, weight=1)

        # Add response_parameters_frame
        self.response_parameters_frame = tk.Frame(self.content_frame)
        self.response_parameters_frame.grid(row=6, column=0, columnspan=3, sticky="w")

        # Response frame
        self.response_frame = tk.Frame(root)
        self.response_frame.grid(row=5, column=0, columnspan=2, padx=5, pady=5, sticky="nsew")

        # Prompt Test Tab
        # Source Directory
        self.source_dir = tk.StringVar()
        self.source_dir_entry = tk.Entry(self.prompt_test_tab, textvariable=self.source_dir, width=50)
        self.source_dir_entry.grid(row=0, column=0, padx=10, pady=10, sticky="w")

        self.source_dir_button = tk.Button(self.prompt_test_tab, text="Source Directory",
                                           command=self.select_source_dir)
        self.source_dir_button.grid(row=0, column=1, padx=10, sticky="w")  # Same row as the Entry, next column

        # Output Directory
        self.output_dir = tk.StringVar()
        self.output_dir_entry = tk.Entry(self.prompt_test_tab, textvariable=self.output_dir, width=50)
        self.output_dir_entry.grid(row=1, column=0, padx=10, pady=10, sticky="w")

        self.output_dir_button = tk.Button(self.prompt_test_tab, text="Output Directory",
                                           command=self.select_output_dir)
        self.output_dir_button.grid(row=1, column=1, padx=10, sticky="w")

        # schema file
        self.schema_file = tk.StringVar()
        self.schema_file_entry = tk.Entry(self.prompt_test_tab, textvariable=self.schema_file, width=50)
        self.schema_file_entry.grid(row=2, column=0, padx=10, pady=10, sticky="w")

        self.schema_file_button = tk.Button(self.prompt_test_tab, text="Schema File",
                                            command=self.select_schema_file)
        self.schema_file_button.grid(row=2, column=1, padx=10, sticky="w")

        # TODO: Add an output file name
        # TODO: Allow the user to specify the delimiters for in the prompt template

        # Test Button
        self.test_button = tk.Button(self.prompt_test_tab, text="Test",
                                     command=self.run_test)
        self.test_button.grid(row=3, columnspan=2, padx=10, pady=10, sticky="w")

        # Token Totals
        self.prompt_token_total = tk.StringVar(value="Prompt Tokens: 0")
        self.completion_token_total = tk.StringVar(value="Completion Tokens: 0")
        self.combined_token_total = tk.StringVar(value="Combined Tokens: 0")

        # Labels to display the token totals
        self.prompt_token_label = tk.Label(self.prompt_test_tab, textvariable=self.prompt_token_total)
        self.prompt_token_label.grid(row=4, column=0, padx=10, pady=10, sticky="w")

        self.completion_token_label = tk.Label(self.prompt_test_tab, textvariable=self.completion_token_total)
        self.completion_token_label.grid(row=5, column=0, padx=10, pady=10, sticky="w")

        self.combined_token_label = tk.Label(self.prompt_test_tab, textvariable=self.combined_token_total)
        self.combined_token_label.grid(row=6, column=0, padx=10, pady=10, sticky="w")

    def select_source_dir(self):
        """Open file dialog to select source directory"""
        source_dir = filedialog.askdirectory()
        self.source_dir.set(source_dir)
        print("Selected source directory:", source_dir)

    def select_output_dir(self):
        """Open file dialog to select output directory"""
        output_dir = filedialog.askdirectory()
        self.output_dir.set(output_dir)
        print("Selected output directory:", output_dir)

    def select_schema_file(self):
        """Open file dialog to select schema file"""
        schema_file = filedialog.askopenfilename()
        self.schema_file.set(schema_file)
        print("Selected schema file:", schema_file)

    def run_test(self):

        # Initialize the schema variable
        schema = None

        print("Running test...")
        try:
            # Load the JSON schema from the file
            with open(self.schema_file.get(), 'r') as schema_f:
                schema = json.load(schema_f)
        except FileNotFoundError:
            print("Error: The schema file was not found.")
            return  # Exit the method if the schema file is not found
        except json.JSONDecodeError:
            print("Error: The schema file does not contain valid JSON.")
            return  # Exit the method if the schema is not valid JSON
        except PermissionError:
            print("Error: Permission denied when trying to read the schema file.")
            return  # Exit the method if there's a permission error

        # If schema is None after the above operations, something went wrong.
        # We should not proceed with the rest of the method.
        if not schema:
            print("Error: Failed to load the JSON schema. Aborting the test run.")
            return

        total_prompt_tokens = 0
        total_completion_tokens = 0
        total_combined_tokens = 0

        with open(os.path.join(self.output_dir.get(), "generated_composite_prompts.txt"),
                  "w") as generated_prompts_file, \
                open(os.path.join(self.output_dir.get(), "raw_responses.txt"), "w") as raw_responses_file:

            for file in os.listdir(self.source_dir.get()):
                print("Processing file:", file)

                # Detect the encoding of the file
                with open(os.path.join(self.source_dir.get(), file), 'rb') as f:
                    result = chardet.detect(f.read())
                    file_encoding = result['encoding']

                # Now read the file with the detected encoding
                with open(os.path.join(self.source_dir.get(), file), 'r', encoding=file_encoding) as f:
                    content = f.read()
                print(f"File content of {file}: {content[:100]}...")  # printing the first 100 characters of content

                # Make a fresh copy of the prompt template for each file
                # test_message_components = copy.deepcopy(self.message_components)

                test_message_components = []

                for msg in self.message_components:
                    print(f"Processing message: {msg}")
                    print(f"msg['role'] type: {type(msg['role'])}, value: {msg['role']}")
                    test_msg = {
                        'role': msg['role'].get() if isinstance(msg['role'], tk.StringVar) else msg['role'],
                        'content': msg['content'].get() if isinstance(msg['content'], tk.StringVar) else msg['content']
                    }
                    print(f"Constructed test message: {test_msg}")
                    test_message_components.append(test_msg)

                # Replace delimiter in test_message_components with file content
                for message in test_message_components:
                    message["content"] = message["content"].replace("--{?}--", content)
                    print(f"Message after delimiter replacement: {message['content'][:100]}...")  # first 100 characters

                # Write generated prompt to file
                generated_prompts_file.write(json.dumps(test_message_components))
                generated_prompts_file.write('\n\n------------------------------\n\n')  # for better readability
                print(f"Written generated prompt for {file}")

                # Check context length
                num_tokens = num_tokens_from_messages(test_message_components, self.model_var.get())
                if num_tokens > self.context_length.get():
                    # Handle too long prompt
                    continue

                # Submit prompt
                try:
                    print(f"Submitting prompt for {file}: {test_message_components}")
                    response = self.submit_test_prompt(test_message_components, self.output_dir.get())

                    print(f"Received response for {file}: {response}")
                    # Write raw response to file
                    raw_responses_file.write(json.dumps(response))
                    raw_responses_file.write('\n\n------------------------------\n\n')  # for better readability

                    # Update token counts
                    if isinstance(response, dict):
                        print(f"Updating total_prompt_tokens count for {file}")
                        total_prompt_tokens += response.get('usage', {}).get("prompt_tokens", 0)
                        print(f"Updating total_completion_tokens count for {file}")
                        total_completion_tokens += response.get('usage', {}).get("completion_tokens", 0)
                        print(f"Updating total_combined_tokens count for {file}")
                        total_combined_tokens += response.get('usage', {}).get("total_tokens", 0)
                    else:
                        print(
                            f"Warning: Unexpected response type for {file}. Expected a dictionary but got {type(response)}.")

                    self.prompt_token_total.set(f"Prompt Tokens: {total_prompt_tokens}")
                    self.completion_token_total.set(f"Completion Tokens: {total_completion_tokens}")
                    self.combined_token_total.set(f"Combined Tokens: {total_combined_tokens}")

                    # Handle JSON extraction and validation
                    if isinstance(response, dict):
                        response_str = json.dumps(response)
                        extracted_json = self.extract_json_from_response(response_str)
                        print(f"JSON extracted from response_str for {file}: {extracted_json}")
                    else:
                        # Log an error message for unexpected response types
                        print(f"Unexpected response type {type(response)} when processing. Expected a dictionary.")
                        # Optionally, raise an exception to halt the process (if this is considered a critical error)
                        # raise TypeError(f"Unexpected response type {type(response)}. Expected a dictionary.")
                    validated_json = self.validate_and_capture_json(extracted_json,
                                                                    schema)
                    print(f"Validated Response JSON for {file}: {validated_json}")

                    # Save the validated JSON with validation details to Excel
                    print(f"Saving validated JSON for {file} in run_test()")
                    output_directory = self.output_dir.get()
                    self.save_expanded_json_to_excel(output_directory, validated_json)

                    work_order_json = None
                    print(f"Extracting work order JSON for {file}")
                    print(f"Choices is in extracted json: ", 'choices' in extracted_json)
                    print(f"Number of choices: ", len(extracted_json['choices']))
                    print("\n\n\n\n\n\n\n\n**************\n\n\n\n\n\n\n\n")
                    if extracted_json and 'choices' in extracted_json and len(extracted_json['choices']) > 0:
                        content_str = extracted_json['choices'][0]['message'].get('content', "")
                        print(f"Content string for {file}: {content_str}")
                        work_order_json = self.extract_json_from_response(content_str)
                        print(f"Work order JSON for {file}: {work_order_json}")

                    # Save the work order to Excel
                    if work_order_json:
                        print(f"Saving work order JSON for {file} in run_test()")
                        self.save_work_order_to_excel(output_directory, work_order_json, schema)

                except Exception as e:
                    # Handle errors during prompt submission, writing to file, or JSON handling
                    print(f"Error processing file {file}: {str(e)}")
                    pass

    def extract_json_from_response(self, response):
        """
        Extracts a JSON object from a string that might have other text.

        Args:
        - response (str): The string containing the JSON object.

        Returns:
        - dict: The extracted JSON object or None if extraction failed.
        """

        start_pos = response.find('{')
        if start_pos == -1:
            return None  # No opening brace found

        for end_pos in range(response.rfind('}') + 1, start_pos, -1):
            # Try to parse the substring as JSON
            try:
                potential_json_str = response[start_pos:end_pos]
                return json.loads(potential_json_str)
            except json.JSONDecodeError:
                continue
        return None

    def validate_and_capture_json(self, data, schema):
        try:
            validate(data, schema)
            expanded_json = {
                "validity_status": True,
                "validation_errors": [],
                "original_json": data,
                "validation_timestamp": datetime.datetime.now().isoformat()
            }
            return expanded_json
        except ValidationError as e:
            error_details = {
                "message": e.message,
                "path": list(e.path),
                "validator": e.validator,
                "validator_value": e.validator_value,
                "instance_value": e.instance
            }
            expanded_json = {
                "validity_status": False,
                "validation_errors": [error_details],
                "original_json": data,
                "validation_timestamp": datetime.datetime.now().isoformat()
            }
            return expanded_json

    def save_expanded_json_to_excel(self, directory, expanded_json):
        file_path = os.path.join(directory, "validation_results.xlsx")

        # If the file doesn't exist, create it
        if not os.path.exists(file_path):
            with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                df_summary = pd.DataFrame(columns=["File Name", "Validation Status", "Timestamp", "Error Count"])
                df_summary.to_excel(writer, sheet_name='Summary', index=False)
                df_originals = pd.DataFrame()
                df_originals.to_excel(writer, sheet_name='Original_JSONs', index=False)
                df_errors = pd.DataFrame()
                df_errors.to_excel(writer, sheet_name='Errors', index=False)

        # Now, process the existing (or newly created) Excel file
        with pd.ExcelWriter(file_path, engine='openpyxl', mode='a') as writer:
            workbook = load_workbook(file_path)
            writer._book = workbook

            # Append to the Summary sheet
            if 'Summary' in workbook.sheetnames:
                df_summary = pd.read_excel(file_path, sheet_name='Summary', engine='openpyxl')
                # Delete the old sheet to write a new updated one
                del writer.book['Summary']
            else:
                df_summary = pd.DataFrame(columns=["File Name", "Validation Status", "Timestamp", "Error Count"])

            new_row = {
                "File Name": "validation_results.xlsx",  # Replace with the actual file name
                "Validation Status": expanded_json["validity_status"],
                "Timestamp": expanded_json["validation_timestamp"],
                "Error Count": len(expanded_json["validation_errors"])
            }
            new_df = pd.DataFrame([new_row])
            df_summary = pd.concat([df_summary, new_df], ignore_index=True)
            df_summary.to_excel(writer, sheet_name='Summary', index=False)

            # Save errors in the same "Errors" sheet
            if 'Errors' in workbook.sheetnames:
                df_errors = pd.read_excel(file_path, sheet_name='Errors', engine='openpyxl')
                del writer.book['Errors']
            else:
                df_errors = pd.DataFrame()

            new_errors_df = pd.DataFrame(expanded_json["validation_errors"])
            df_errors = pd.concat([df_errors, new_errors_df], ignore_index=True)
            df_errors.to_excel(writer, sheet_name='Errors', index=False)

            # Save the original JSON in a flattened format
            if 'Original_JSONs' in workbook.sheetnames:
                df_originals = pd.read_excel(file_path, sheet_name='Original_JSONs', engine='openpyxl')
                del writer.book['Original_JSONs']
            else:
                df_originals = pd.DataFrame()

            flattened_json = pd.json_normalize(expanded_json["original_json"])
            df_originals = pd.concat([df_originals, flattened_json], ignore_index=True)
            df_originals.to_excel(writer, sheet_name='Original_JSONs', index=False)

    def save_work_order_to_excel(self, directory, work_order_json, schema):
        file_path = os.path.join(directory, "work_orders.xlsx")

        # Convert JSON to a DataFrame for the main data
        df_main = pd.json_normalize(work_order_json)
        print(f"df_main:  \n", df_main)

        # Handle arrays based on schema
        for key, value in schema["properties"].items():
            if value["type"] == "array":
                max_items = value["maxItems"]
                item_props = value["items"]["properties"]

                for i in range(max_items):
                    for prop, prop_details in item_props.items():
                        col_name = f"{key}_{i + 1}_{prop}"
                        try:
                            df_main[col_name] = work_order_json[key][i][prop]
                        except (IndexError, KeyError):
                            df_main[col_name] = None

        # If the Excel file doesn't exist, create it
        if not os.path.exists(file_path):
            with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                df_main.to_excel(writer, sheet_name="Work Orders", index=False)
        else:
            # If the Excel file exists, append the new work order data to the existing data
            with pd.ExcelWriter(file_path, engine='openpyxl', mode='a') as writer:
                workbook = load_workbook(file_path)
                writer._book = workbook

                # If the "Work Orders" sheet exists, load the existing data and append the new data
                if "Work Orders" in workbook.sheetnames:
                    df_existing = pd.read_excel(file_path, sheet_name='Work Orders', engine='openpyxl')
                    df_combined = pd.concat([df_existing, df_main], ignore_index=True)
                    # Delete the old sheet to write the updated one
                    del writer.book["Work Orders"]
                    df_combined.to_excel(writer, sheet_name="Work Orders", index=False)
                else:
                    # If the "Work Orders" sheet doesn't exist, simply write the new work order data
                    df_main.to_excel(writer, sheet_name="Work Orders", index=False)

    def submit_test_prompt(self, test_message_components, output_dir):
        """
        The declaration for the send request method is: def send_request(model, prompt, temperature, top_p, n, stream,
        stop, max_tokens, presence_penalty, frequency_penalty, logit_bias, user):
        """

        # Validate model based on the available models supported by the API
        """
        Valid values: gpt-4, gpt-3.5-turbo, gpt-3.5-turbo-0301, gpt-4-0314, gpt-4-32k (not yet available)
        """
        print("Validating parameters and submitting prompt to LLM API")
        if not self.validate_model():
            return
        print(f"Attempting to get the model value.")
        model = self.model_var.get()
        print(f"Got the model value.")

        # Assemble the prompt and validate the total number of tokens
        messages_list = []
        """
        Required. Messages is an array of messages in the format "[role]":"[message]".
        """
        for item in test_message_components:
            # role = item['role'].get()
            role = item['role']
            content = item['content']
            messages_list.append({'role': role, 'content': content})

        # Validate the total number of tokens in the prompt
        total_tokens = num_tokens_from_messages(messages_list, model=self.model_var.get())
        context_length = self.context_windows[model]
        print(f"Total number of tokens: {total_tokens}")
        print(f"Context length: {context_length}")
        if total_tokens > context_length:
            messagebox.showerror("Error", f"The total number of tokens in the prompt exceeds {context_length}. "
                                          f"Please reduce the length of the prompt.")

            print(f"Total number of tokens in the message set is too great: {total_tokens}")
            return
        else:
            prompt = messages_list
            print(f"Constructed prompt: {prompt}")

        # Validate temperature between 0 and 2
        """
        Optional. Temperature is a number that defaults to 1. It indicates what sampling temperature to use, between
        0 and 2. Higher values like 0.8 will make the output more random, while lower values like 0.2 will make it
        more focused and deterministic. It is generally recommend altering this or top_p but not both.
        """
        try:
            temperature = float(self.temperature_entry.get())
            if temperature < 0 or temperature > 2:
                raise ValueError("Invalid temperature value")
        except ValueError:
            message_box = CustomMessageBox(
                self.parameters_frame, "Invalid temperature value",
                "Invalid temperature value entered. Continue with the default value (0.7) or return to the UI to"
                " provide a valid value?"
            )
            if message_box.result:
                temperature = 0.7
            else:
                return

        # Validate top_p between 0 and 1
        """
        Optional. Top_p is a number that defaults to 1. It is an alternative to sampling with temperature, called
        nucleus sampling, where the model considers the results of the tokens with top_p probability mass. So 0.1 means
        only the tokens comprising the top 10% probability mass are considered. We generally recommend altering this or
        temperature but not both.
        """
        # Validate top_p between 0 and 1
        try:
            top_p = float(self.top_p_entry.get())
            if top_p < 0 or top_p > 1:
                raise ValueError("Invalid top_p value")
        except ValueError:
            message_box = CustomMessageBox(
                self.parameters_frame, "Invalid top_p value",
                "Invalid top_p value entered. Continue with the default value (1) or return to the main window to "
                "provide a valid value?"
            )
            if message_box.result:
                top_p = 1
            else:
                return

        # Validate n between 1 and 4
        """
        Optional. N is an integer between 1 and 4 that defaults to 1. It controls the number of completions to
        generate. If n is greater than 1, the API will return a list of completions. If the value is set to 1, only
        the best completion will be returned.
        """
        try:
            n = int(self.n_entry.get())
            if n < 1 or n > 4:
                raise ValueError("Invalid n value")
        except ValueError:
            message_box = CustomMessageBox(
                self.parameters_frame, "Invalid n value",
                "Invalid n value entered. Continue with the default value (1) or return to the main window to provide"
                " a valid value?"
            )
            if message_box.result:
                n = 1
            else:
                return

        # Set the stream to False
        """
        Optional. Stream is a boolean that defaults to false. If set, partial message deltas will be sent, like in
        ChatGPT. Tokens will be sent as data-only server-sent events as they become available, with the stream
        terminated by a data: [DONE] message.
        """
        stream = False  # self.stream_entry.get()

        # Validate stop
        """
        Optional. Stop is a string or array of up to 4 sequences where the API will stop generating further tokens.
        """
        try:
            stop = self.stop_entry.get()
            if stop:
                # Remove leading/trailing spaces and quotation marks
                stop = [s.strip().strip('\'"') for s in stop.split(',')]
                if len(stop) > 4:
                    raise ValueError("Too many stop sequences")
        except ValueError:
            message_box = CustomMessageBox(
                self.parameters_frame, "Invalid stop value",
                "Invalid stop value entered. Continue with the default value (empty) or return to the main window to"
                " provide a valid value?"
            )
            if message_box.result:
                stop = []
            else:
                return

        # Validate max_tokens based on the number of tokens in the prompt
        """
        Optional. Max_tokens is an integer that defaults to infinite. It indicates the maximum number of tokens the
        model is to generate in the chat completion. The total length of input tokens and generated tokens is limited
        by the model's context length. For GPT-4, the context length is 4096 tokens.
        """
        # TODO: Validate max_tokens based on the number of tokens in the context window for the selected model
        try:
            max_tokens_str = self.max_tokens_entry.get()
            if not max_tokens_str:
                max_tokens = None
            else:
                max_tokens = int(max_tokens_str)
                if max_tokens < 1 or max_tokens > self.context_windows:
                    raise ValueError("Invalid max_tokens value")
        except ValueError:
            message_box = CustomMessageBox(
                self.parameters_frame, "Invalid max_tokens value",
                "Invalid max_tokens value entered. Continue with the default value (infinite) or return to the main"
                " window to provide a valid value?"
            )
            if message_box.result:
                max_tokens = None
            else:
                return

        # Validate presence_penalty
        """
        Optional. Presence_penalty is a number between -2.0 and 2.0 that defaults to 0. Positive values penalize new
        tokens based on whether they appear in the text so far, increasing the model's likelihood to talk about new
        topics.
        """
        try:
            presence_penalty_str = self.presence_penalty_entry.get()
            if not presence_penalty_str:
                presence_penalty = 0
            else:
                presence_penalty = float(presence_penalty_str)
                if presence_penalty < -2.0 or presence_penalty > 2.0:
                    raise ValueError("Invalid presence_penalty value")
        except ValueError:
            message_box = CustomMessageBox(
                self.parameters_frame, "Invalid presence_penalty value",
                "Invalid presence_penalty value entered. Continue with the default value (0) or return to the main"
                " window to provide a valid value?"
            )
            if message_box.result:
                presence_penalty = 0
            else:
                return

        # Validate frequency_penalty
        """
        Optional. Frequency_penalty is a number between -2.0 and 2.0 that defaults to 0. Positive values penalize new
        tokens based on their existing frequency in the text so far, decreasing the model's likelihood to repeat the
        same line verbatim.
        """
        try:
            frequency_penalty_str = self.frequency_penalty_entry.get()
            if not frequency_penalty_str:
                frequency_penalty = 0
            else:
                frequency_penalty = float(frequency_penalty_str)
                if frequency_penalty < -2.0 or frequency_penalty > 2.0:
                    raise ValueError("Invalid frequency_penalty value")
        except ValueError:
            message_box = CustomMessageBox(
                self.parameters_frame, "Invalid frequency_penalty value",
                "Invalid frequency_penalty value entered. Continue with the default value (0) or return to the main"
                " window to provide a valid value?"
            )
            if message_box.result:
                frequency_penalty = 0
            else:
                return

        # Get the logit_bias value and parse it into a dictionary
        """
        Optional. Logit_bias is a map that defaults to null. It modifies the likelihood of specified tokens appearing
        in the completion. It accepts a json object that maps tokens (specified by their token ID in the tokenizer) to
        an associated bias value from -100 to 100. Mathematically, the bias is added to the logits generated by the
        model prior to sampling. The exact effect will vary per model, but values between -1 and 1 should decrease or
        increase likelihood of selection; values like -100 or 100 should result in a ban or exclusive selection of
        the relevant token.
        """
        logit_bias_str = self.logit_bias_entry.get()
        if not logit_bias_str.strip():  # Treat an empty value as acceptable
            logit_bias = {}
        else:
            try:
                logit_bias = json.loads(logit_bias_str)
            except json.JSONDecodeError:
                print("Invalid JSON: ", logit_bias_str)
                logit_bias = {}  # Use an empty dictionary if the value is not valid JSON

        # Get the user value
        """
        Optional. User is a string that defaults to null. The length should not exceed 256 characters. User is a unique
        identifier representing your end-user, which can help OpenAI to monitor and detect abuse.
        """
        user = self.user_entry.get()
        if len(user) > 256:
            message_box = CustomMessageBox(
                self.parameters_frame, "Invalid user value",
                "User value exceeds 256 characters. Continue with the default value (null) or return to the main window"
                " to provide a valid value?"
            )
            if message_box.result:
                user = None
            else:
                return

        # (model, prompt, temperature, top_p, n, stop, max_tokens, presence_penalty, frequency_penalty, logit_bias,
        #  user)
        print("Parameters validated. Sending request...")

        # TODO: Add a delimiter between the requests for each file
        # Add the elements that will compose the request to the request file in the output directory
        request = {
            "model": model,
            "prompt": prompt,
            "temperature": temperature,
            "top_p": top_p,
            "n": n,
            "stream": stream,
            "stop": stop,
            "max_tokens": max_tokens,
            "presence_penalty": presence_penalty,
            "frequency_penalty": frequency_penalty,
            "logit_bias": logit_bias,
            "user": user
        }
        with open(os.path.join(output_dir, "request.json"), "w") as request_file:
            json.dump(request, request_file, indent=4)

        print("Sending request with the following parameters:")
        print(f"Model: {model}")
        print(f"Prompt: {prompt}")
        print(f"Temperature: {temperature}")
        print(f"Top_p: {top_p}")
        print(f"N: {n}")
        print(f"Stream: {stream}")
        print(f"Stop: {stop}")
        print(f"Max tokens: {max_tokens}")
        print(f"Presence penalty: {presence_penalty}")
        print(f"Frequency penalty: {frequency_penalty}")
        print(f"Logit bias: {logit_bias}")
        print(f"User: {user}")
        # Send the request
        response = send_request(model, prompt, temperature, top_p, n, stream, stop, max_tokens, presence_penalty,
                                frequency_penalty, logit_bias, user)
        print(f"Received response: {response}")
        self.display_response(response)
        return response

    def update_context_length(self, *args):
        selected_model = self.model_var.get()
        print("Updating context window value for selected model: " + selected_model)
        if selected_model in self.context_windows:
            context_length = self.context_windows[selected_model]
        else:
            context_length = 0
        self.context_length.set(context_length)

    def refresh_model_list(self):
        new_model_list = populate_model_list(self.context_windows, self.api_key)
        print("Updating model list: " + str(new_model_list))
        self.model_menu['values'] = ["Select Model"] + new_model_list
        self.model_list = new_model_list

    def edit_api_key(self):
        # Here is where you'd put the logic to edit the API Key
        new_api_key = get_api_key(self.root)
        if new_api_key:
            self.api_key_entry.delete(0, 'end')
            self.api_key_entry.insert(0, new_api_key)

    def validate_model(self):

        # Validate model based on the available models supported by the API
        """
        Required. Model is a string containing the ID of the model to use.
        Valid values: gpt-4, gpt-3.5-turbo, gpt-3.5-turbo-0301, gpt-4-0314, gpt-4-32k (not yet available)
        """
        print("Validating model")
        model = self.model_var.get()
        if model == "Refresh Model List":
            populate_model_list(self.context_windows, self.model_list)
            return False
        if model not in self.model_list:
            messagebox.showerror("Error", "Invalid model. Please select a valid model.")
            print(f"Model is invalid", {model})
            return False
        print(f"Model is valid", {model})
        return True

    # Method to open the message component window
    def open_message_component_window(self, content=None):

        """
        Valid values: gpt-4, gpt-3.5-turbo, gpt-3.5-turbo-0301, gpt-4-0314, gpt-4-32k (not yet available)
        """
        if not self.validate_model():
            return

        message_component_window = tk.Toplevel(self.root)
        message_component_window.geometry("800x800")
        text_box = ScrolledText(message_component_window, wrap=tk.WORD)
        text_box.pack(padx=10, pady=10, fill=tk.BOTH, expand=True)

        print(f"Opening message component window")
        if content:
            text_box.insert("1.0", content)

        save_button = tk.Button(message_component_window, text="Save", command=lambda:
        self.save_message_component(text_box.get("1.0", "end-1c"), message_component_window))
        save_button.pack(pady=10)

    # Method to save the message component and close the message component window
    def save_message_component(self, content, window):
        print("Saving message component")
        self.add_message_component(content)
        self.update_message_component_list()
        window.destroy()

    # Method to add message component to the list and update the token count
    def add_message_component(self, content):
        print("Adding message component")
        self.message_components.append({'role': tk.StringVar(), 'content': content})

        # Include all the message components in the token count
        self.token_count.set(num_tokens_from_messages(self.message_components, model=self.model_var.get()))

    # Method to update the message component list in the main window
    def update_message_component_list(self):
        print("Updating message component list")
        for widget in self.message_components_frame.winfo_children():
            widget.destroy()

        for index, message_item in enumerate(self.message_components):
            row_frame = tk.Frame(self.message_components_frame)
            row_frame.pack(fill=tk.X, padx=5, pady=5)

            role_menu = ttk.Combobox(row_frame, textvariable=message_item['role'],
                                     values=["system", "user", "assistant"], width=10)
            role_menu.set(message_item['role'].get() if message_item['role'].get() else "user")
            role_menu.pack(side=tk.LEFT, padx=5)

            displayed_content = message_item['content'][:50] + " . . . " if len(message_item['content']) > 50 \
                else message_item['content']
            content_label = tk.Label(row_frame, text=displayed_content, width=60, anchor='w')
            content_label.pack(side=tk.LEFT, padx=5)

            edit_button = tk.Button(row_frame, text="Edit", command=lambda item=message_item:
            self.edit_message_component(item))
            edit_button.pack(side=tk.RIGHT, padx=10)

            up_button = tk.Button(row_frame, text="↑", command=lambda i=index:
            self.move_message_component_up(i))
            up_button.pack(side=tk.RIGHT, padx=(0, 2))
            if index == 0:  # If it's the first item
                up_button["state"] = "disabled"

            down_button = tk.Button(row_frame, text="↓", command=lambda i=index:
            self.move_message_component_down(i))
            down_button.pack(side=tk.RIGHT, padx=(2, 0))
            if index == len(self.message_components) - 1:  # If it's the last item
                down_button["state"] = "disabled"

    def move_message_component_up(self, index):
        print("Moving message component up")
        if index > 0:  # No need to move if it's the first item
            self.message_components[index], self.message_components[index - 1] = self.message_components[index - 1], \
                self.message_components[index]
            self.update_message_component_list()

    def move_message_component_down(self, index):
        print("Moving message component down")
        if index < len(self.message_components) - 1:  # No need to move if it's the last item
            self.message_components[index], self.message_components[index + 1] = self.message_components[index + 1], \
                self.message_components[index]
            self.update_message_component_list()

    # Method to edit the message component
    def edit_message_component(self, item):
        def save_changes():
            print(f"Saving changes to message component")
            item['content'] = text_box.get("1.0", "end-1c")
            self.update_message_component_list()

            # Include all the message components in the token count
            self.token_count.set(num_tokens_from_messages(self.message_components, model=self.model_var.get()))

            message_component_window.destroy()

        print(f"Editing message component")

        message_component_window = tk.Toplevel(self.root)
        message_component_window.geometry("800x800")
        text_box = ScrolledText(message_component_window, wrap=tk.WORD)
        text_box.pack(padx=10, pady=10, fill=tk.BOTH, expand=True)

        text_box.insert("1.0", item['content'])

        save_button = tk.Button(message_component_window, text="Save", command=save_changes)
        save_button.pack(pady=10)

    # Method to handle the submit prompt button click event
    def submit_prompt(self):
        """
        The declaration for the send request method is: def send_request(model, prompt, temperature, top_p, n, stream,
        stop, max_tokens, presence_penalty, frequency_penalty, logit_bias, user):
        """

        # Validate model based on the available models supported by the API
        """
        Valid values: gpt-4, gpt-3.5-turbo, gpt-3.5-turbo-0301, gpt-4-0314, gpt-4-32k (not yet available)
        """
        print("Validating parameters and submitting prompt to LLM API")
        if not self.validate_model():
            return
        print(f"Attempting to get the model value.")
        model = self.model_var.get()

        # Assemble the prompt and validate the total number of tokens
        messages_list = []
        """
        Required. Messages is an array of messages in the format "[role]":"[message]".
        """
        for item in self.message_components:
            role = item['role']
            content = item['content']
            messages_list.append({'role': role, 'content': content})

        # Validate the total number of tokens in the prompt
        total_tokens = num_tokens_from_messages(messages_list, model=self.model_var.get())
        context_length = self.context_windows[model]
        print(f"Total number of tokens: {total_tokens}")
        print(f"Context length: {context_length}")
        if total_tokens > context_length:
            messagebox.showerror("Error", f"The total number of tokens in the prompt exceeds {context_length}. "
                                          f"Please reduce the length of the prompt.")

            print(f"Total number of tokens in the message set is too great: {total_tokens}")
            return
        else:
            prompt = messages_list

        # Validate temperature between 0 and 2
        """
        Optional. Temperature is a number that defaults to 1. It indicates what sampling temperature to use, between
        0 and 2. Higher values like 0.8 will make the output more random, while lower values like 0.2 will make it
        more focused and deterministic. It is generally recommend altering this or top_p but not both.
        """
        try:
            temperature = float(self.temperature_entry.get())
            if temperature < 0 or temperature > 2:
                raise ValueError("Invalid temperature value")
        except ValueError:
            message_box = CustomMessageBox(
                self.parameters_frame, "Invalid temperature value",
                "Invalid temperature value entered. Continue with the default value (0.7) or return to the UI to"
                " provide a valid value?"
            )
            if message_box.result:
                temperature = 0.7
            else:
                return

        # Validate top_p between 0 and 1
        """
        Optional. Top_p is a number that defaults to 1. It is an alternative to sampling with temperature, called
        nucleus sampling, where the model considers the results of the tokens with top_p probability mass. So 0.1 means
        only the tokens comprising the top 10% probability mass are considered. We generally recommend altering this or
        temperature but not both.
        """
        # Validate top_p between 0 and 1
        try:
            top_p = float(self.top_p_entry.get())
            if top_p < 0 or top_p > 1:
                raise ValueError("Invalid top_p value")
        except ValueError:
            message_box = CustomMessageBox(
                self.parameters_frame, "Invalid top_p value",
                "Invalid top_p value entered. Continue with the default value (1) or return to the main window to "
                "provide a valid value?"
            )
            if message_box.result:
                top_p = 1
            else:
                return

        # Validate n between 1 and 4
        """
        Optional. N is an integer between 1 and 4 that defaults to 1. It controls the number of completions to
        generate. If n is greater than 1, the API will return a list of completions. If the value is set to 1, only
        the best completion will be returned.
        """
        try:
            n = int(self.n_entry.get())
            if n < 1 or n > 4:
                raise ValueError("Invalid n value")
        except ValueError:
            message_box = CustomMessageBox(
                self.parameters_frame, "Invalid n value",
                "Invalid n value entered. Continue with the default value (1) or return to the main window to provide"
                " a valid value?"
            )
            if message_box.result:
                n = 1
            else:
                return

        # Set the stream to False
        """
        Optional. Stream is a boolean that defaults to false. If set, partial message deltas will be sent, like in
        ChatGPT. Tokens will be sent as data-only server-sent events as they become available, with the stream
        terminated by a data: [DONE] message.
        """
        stream = False  # self.stream_entry.get()

        # Validate stop
        """
        Optional. Stop is a string or array of up to 4 sequences where the API will stop generating further tokens.
        """
        try:
            stop = self.stop_entry.get()
            if stop:
                # Remove leading/trailing spaces and quotation marks
                stop = [s.strip().strip('\'"') for s in stop.split(',')]
                if len(stop) > 4:
                    raise ValueError("Too many stop sequences")
        except ValueError:
            message_box = CustomMessageBox(
                self.parameters_frame, "Invalid stop value",
                "Invalid stop value entered. Continue with the default value (empty) or return to the main window to"
                " provide a valid value?"
            )
            if message_box.result:
                stop = []
            else:
                return

        # Validate max_tokens based on the number of tokens in the prompt
        """
        Optional. Max_tokens is an integer that defaults to infinite. It indicates the maximum number of tokens the
        model is to generate in the chat completion. The total length of input tokens and generated tokens is limited
        by the model's context length. For GPT-4, the context length is 4096 tokens.
        """
        # TODO: Validate max_tokens based on the number of tokens in the context window for the selected model
        try:
            max_tokens_str = self.max_tokens_entry.get()
            if not max_tokens_str:
                max_tokens = None
            else:
                max_tokens = int(max_tokens_str)
                if max_tokens < 1 or max_tokens > self.context_windows:
                    raise ValueError("Invalid max_tokens value")
        except ValueError:
            message_box = CustomMessageBox(
                self.parameters_frame, "Invalid max_tokens value",
                "Invalid max_tokens value entered. Continue with the default value (infinite) or return to the main"
                " window to provide a valid value?"
            )
            if message_box.result:
                max_tokens = None
            else:
                return

        # Validate presence_penalty
        """
        Optional. Presence_penalty is a number between -2.0 and 2.0 that defaults to 0. Positive values penalize new
        tokens based on whether they appear in the text so far, increasing the model's likelihood to talk about new
        topics.
        """
        try:
            presence_penalty_str = self.presence_penalty_entry.get()
            if not presence_penalty_str:
                presence_penalty = 0
            else:
                presence_penalty = float(presence_penalty_str)
                if presence_penalty < -2.0 or presence_penalty > 2.0:
                    raise ValueError("Invalid presence_penalty value")
        except ValueError:
            message_box = CustomMessageBox(
                self.parameters_frame, "Invalid presence_penalty value",
                "Invalid presence_penalty value entered. Continue with the default value (0) or return to the main"
                " window to provide a valid value?"
            )
            if message_box.result:
                presence_penalty = 0
            else:
                return

        # Validate frequency_penalty
        """
        Optional. Frequency_penalty is a number between -2.0 and 2.0 that defaults to 0. Positive values penalize new
        tokens based on their existing frequency in the text so far, decreasing the model's likelihood to repeat the
        same line verbatim.
        """
        try:
            frequency_penalty_str = self.frequency_penalty_entry.get()
            if not frequency_penalty_str:
                frequency_penalty = 0
            else:
                frequency_penalty = float(frequency_penalty_str)
                if frequency_penalty < -2.0 or frequency_penalty > 2.0:
                    raise ValueError("Invalid frequency_penalty value")
        except ValueError:
            message_box = CustomMessageBox(
                self.parameters_frame, "Invalid frequency_penalty value",
                "Invalid frequency_penalty value entered. Continue with the default value (0) or return to the main"
                " window to provide a valid value?"
            )
            if message_box.result:
                frequency_penalty = 0
            else:
                return

        # Get the logit_bias value and parse it into a dictionary
        """
        Optional. Logit_bias is a map that defaults to null. It modifies the likelihood of specified tokens appearing
        in the completion. It accepts a json object that maps tokens (specified by their token ID in the tokenizer) to
        an associated bias value from -100 to 100. Mathematically, the bias is added to the logits generated by the
        model prior to sampling. The exact effect will vary per model, but values between -1 and 1 should decrease or
        increase likelihood of selection; values like -100 or 100 should result in a ban or exclusive selection of
        the relevant token.
        """
        logit_bias_str = self.logit_bias_entry.get()
        if not logit_bias_str.strip():  # Treat an empty value as acceptable
            logit_bias = {}
        else:
            try:
                logit_bias = json.loads(logit_bias_str)
            except json.JSONDecodeError:
                print("Invalid JSON: ", logit_bias_str)
                logit_bias = {}  # Use an empty dictionary if the value is not valid JSON

        # Get the user value
        """
        Optional. User is a string that defaults to null. The length should not exceed 256 characters. User is a unique
        identifier representing your end-user, which can help OpenAI to monitor and detect abuse.
        """
        user = self.user_entry.get()
        if len(user) > 256:
            message_box = CustomMessageBox(
                self.parameters_frame, "Invalid user value",
                "User value exceeds 256 characters. Continue with the default value (null) or return to the main window"
                " to provide a valid value?"
            )
            if message_box.result:
                user = None
            else:
                return

        # (model, prompt, temperature, top_p, n, stop, max_tokens, presence_penalty, frequency_penalty, logit_bias,
        #  user)
        print("Parameters validated. Sending request...")
        response = send_request(model, prompt, temperature, top_p, n, stream, stop, max_tokens, presence_penalty,
                                frequency_penalty, logit_bias, user)
        self.display_response(response)

    def reset_prompt_gui(self):
        print(
            "reset_prompt_gui() called")  # Add this line to check if the method is called when you click the button
        print(self.message_components)  # Print out what's inside self.message_components

    def restore_default_values(self):
        # Reset entries
        self.model_var.set("Select Model")
        self.api_key_entry.delete(0, tk.END)
        self.api_key_entry.insert(0, openai.api_key)
        self.temperature_entry.delete(0, tk.END)
        self.temperature_entry.insert(0, "0.7")
        self.top_p_entry.delete(0, tk.END)
        self.top_p_entry.insert(0, "1.0")
        self.n_entry.delete(0, tk.END)
        self.n_entry.insert(0, "1")
        self.stop_entry.delete(0, tk.END)
        self.max_tokens_entry.delete(0, tk.END)
        self.presence_penalty_entry.delete(0, tk.END)
        self.frequency_penalty_entry.delete(0, tk.END)
        self.logit_bias_entry.delete(0, tk.END)
        self.user_entry.delete(0, tk.END)

    def delete_prompt(self):
        print("delete_prompt() called")

        # Clear the list of message components
        self.message_components.clear()

        # Clear the widgets in the message_components_frame
        for widget in self.message_components_frame.winfo_children():
            widget.destroy()

        # Clear the widgets in the response_parameters_frame
        for widget in self.response_parameters_frame.winfo_children():
            widget.destroy()

        # Clear the results_tab
        for widget in self.results_tab.winfo_children():
            widget.destroy()
        response_text = ScrolledText(self.results_tab, wrap=tk.WORD)
        response_text.configure(state='disabled')
        response_text.pack(padx=10, pady=10, fill=tk.BOTH, expand=True)

        # Reset token count
        self.token_count.set(0)

    def display_response(self, response):
        print("Response received. Displaying response...")
        for widget in self.results_tab.winfo_children():
            widget.destroy()

        # Clear response_parameters_frame
        for widget in self.response_parameters_frame.winfo_children():
            widget.destroy()

        finish_reason = response.choices[0].finish_reason

        # Blank row
        blank_row = tk.Label(self.response_parameters_frame, text="")
        blank_row.grid(row=0, column=0, padx=5, pady=2, sticky="w")

        # Show response parameters in response_parameters_frame
        response_id_label = tk.Label(self.response_parameters_frame, text=f"Response ID: {response.id}")
        response_id_label.grid(row=1, column=0, padx=5, pady=2, sticky="w")

        created_label = tk.Label(self.response_parameters_frame, text=f"Response created on: {response.created}")
        created_label.grid(row=2, column=0, padx=5, pady=2, sticky="w")

        # Format the usage information
        # Usage label
        usage_label = tk.Label(self.response_parameters_frame, text="Usage:")
        usage_label.grid(row=3, column=0, padx=5, pady=2, sticky="w")

        # Usage text
        usage_text = f"Completion tokens: {response.usage['completion_tokens']}\n" \
                     f"Prompt tokens: {response.usage['prompt_tokens']}\n" \
                     f"Total tokens: {response.usage['total_tokens']}"
        usage_text_widget = tk.Text(self.response_parameters_frame, wrap=tk.WORD, height=4, width=30)
        usage_text_widget.insert(tk.END, usage_text)
        usage_text_widget.configure(state='disabled')
        usage_text_widget.grid(row=4, column=0, padx=5, pady=2, sticky="w")

        finish_reason_label = tk.Label(self.response_parameters_frame, text=f"Finish Reason: {finish_reason}")
        finish_reason_label.grid(row=5, column=0, columnspan=2, padx=5, pady=2, sticky="w")

        note_label = tk.Label(text="The response message can be viewed on the Response Tab.", font=("Helvetica", 10,
                                                                                                    "italic", "bold"))
        note_label.grid(row=6, column=0, columnspan=2, padx=5, pady=2, sticky="nw")

        response_text = ScrolledText(self.results_tab, wrap=tk.WORD)
        for i, choice in enumerate(response.choices, start=1):
            response_text.insert(tk.END, f"Response {i}:\n\"{choice.message.content}\"\n\n")
        response_text.configure(state='disabled')
        response_text.pack(padx=10, pady=10, fill=tk.BOTH, expand=True)


def populate_model_list(context_window, api_key):
    """Iterates through the set of models in the context_window global variable and tries an api call for that model
     using the is_valid_api_key_model function. If the API call is successful, adds the model name to the model_list,
     otherwise proceeds to the next model in the context_window list. When all models have been tested, returns the
     model_list.

  Args:
    api_key: The OpenAI API key to use for the API calls.
    context_window: A dictionary of model names and context window sizes.

  Returns:
    A list of the model names that are valid for the provided API key.
  """
    # TODO: Dynamically generate the context_window dictionary from JSON returned from this API call:
    # models = openai.Model.list()
    # for model in models['data']:
    #     print(model['id'])
    #
    # or somewhere like this:
    # https://github.com/terminalcommandnewsletter/everything-chatgpt
    model_list = []
    for model_name, context_window in context_window.items():
        print(f"Testing {model_name}...")
        if not is_valid_api_key_model(api_key, model_name):
            model_list.append(model_name)

    return model_list


if __name__ == "__main__":
    context_windows = {
        "gpt-4": 8192,
        "gpt-4-0613": 8192,
        "gpt-4-0314": 8192,
        "gpt-4-32k": 32768,
        "gpt-4-32k-0613": 32768,
        "gpt-4-32k-0314": 32768,
        "gpt-3.5-turbo": 4096,
        "gpt-3.5-turbo-16k": 16384,
        "gpt-3.5-turbo-0613": 4096,
        "gpt-3.5-turbo-instruct": 4096,
        "gpt-3.5-turbo-0301": 4096,
        "gpt-3.5-turbo-16k-0613": 16384
    }
    api_key = None
    app_root = tk.Tk()
    openai.api_key = check_api_key(app_root, api_key)
    model_list = populate_model_list(context_windows, openai.api_key)
    app = PromptUI(app_root, model_list, context_windows, openai.api_key)
    app_root.mainloop()
